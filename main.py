import math
import tkinter
from math import sqrt
import Levenshtein
import numpy
import win32api
from tkinter import *
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference, PieChart, AreaChart, BarChart, RadarChart
from openpyxl.styles import Border, Side
from py2neo import Graph

from DataBaseQuery.SearchByCyther import NEO4J_DB
from pyecharts.charts import Sankey, WordCloud

from download_main import run_crawl_process
from import_main import getAllTxtData

# 学科列表
basicAndTechTbl = [
('AGRICULTURAL SCIENCES', '农业科学', 'basic'),
('BIOLOGY & BIOCHEMISTRY', '生物学和生物化学', 'basic'),
('CHEMISTRY', '化学', 'basic'),
('CLINICAL MEDICINE', '临床医学', 'basic'),
('COMPUTER SCIENCE', '计算机科学', 'tech'),
('ECONOMICS & BUSINESS', '经济与商学', 'basic'),
('ENGINEERING', '工程', 'tech'),
('ENVIRONMENT/ECOLOGY', '环境/生态学', 'basic'),
('GEOSCIENCES', '地球科学', 'basic'),
('IMMUNOLOGY', '免疫学', 'basic'),
('MATERIALS SCIENCE', '材料科学', 'tech'),
('MATHEMATICS', '数学', 'basic'),
('MICROBIOLOGY', '微生物学', 'basic'),
('MOLECULAR BIOLOGY & GENETICS', '分子生物学与遗传学', 'basic'),
('Multidisciplinary', '多学科', 'basic'),
('NEUROSCIENCE & BEHAVIOR', '神经科学与行为', 'basic'),
('PHARMACOLOGY & TOXICOLOGY', '药理学与毒理学', 'basic'),
('PHYSICS', '物理学', 'basic'),
('PLANT & ANIMAL SCIENCE', '动植物科学', 'basic'),
('PSYCHIATRY/PSYCHOLOGY', '精神病学/心理学', 'basic'),
('SOCIAL SCIENCES, GENERAL', '社会科学概论', 'basic'),
('SPACE SCIENCE', '空间科学 ', 'basic')

]
# 国家列表
nationList = (
            'USA',
            'Abkhazia',
            'Afghanistan',
            'Albania',
            'Algeria',
            'Andorra',
            'Angola',
            'Antigua and Barbuda',
            'Argentina',
            'Armenia',
            'Australia',
            'Austria',
            'Azerbaijan',
            'Commonwealth oftheBahamas',
            'Bahrain',
            'Bangladesh',
            'Barbados',
            'Belarus',
            'Belgium',
            'Belize',
            'Benin',
            'Bhutan',
            'Bolivia',
            'Bosnia and Herzegovina',
            'Botswana',
            'Brazil',
            'Brunei',
            'Bulgaria',
            'Burkina Faso',
            'BurundiCambodia',
            'Cameroon',
            'Canada',
            'Cape Verde',
            'Catalen',
            'Central African Republic',
            'Chad',
            'Chile',
            'China',
            'Colombia',
            'Comoros',
            'Congo (Brazzaville)',
            'Congo (Kinshasa)',
            'Cook Islands',
            'Costa Rica',
            'Côte d\'Ivoire',
            'Croatia',
            'Cuba',
            'Cyprus',
            'Czech RepublicDenmark',
            'Djibouti',
            'Donetsk People\'s Republic',
            'Dominica',
            'Dominican Republic',
            'Ecuador',
            'Egypt',
            'El Salvador',
            'England',
            'Equatorial Guinea',
            'Eritrea',
            'Estonia',
            'Ethiopia',
            'Fiji',
            'Finland',
            'France',
            'Gabon',
            'Gambia',
            'Georgia',
            'Germany',
            'Ghana',
            'Greece',
            'Grenada',
            'Guatemala',
            'Guinea',
            'Guinea-Bissau',
            'Guyana',
            'Haiti',
            'Honduras',
            'Hungary',
            'Iceland',
            'India',
            'Indonesia',
            'Iran',
            'Iraq',
            'Ireland',
            'Israel',
            'Italy',
            'Jamaica',
            'Japan',
            'Jordan',
            'Kazakhstan',
            'Kenya',
            'Kiribati',
            'South Korea',
            'Kosovo',
            'Kuwait',
            'Kyrgyzstan',
            'Laos',
            'Latvia',
            'Lebanon',
            'Lesotho',
            'Liberia',
            'Libya',
            'Liechtenstein',
            'Lithuania',
            'Luxembourg',
            'Madagascar',
            'Malawi',
            'Malaysia',
            'Maldives',
            'Maltese Knights',
            'Mali',
            'Malta',
            'Marshall Islands',
            'Mauritania',
            'Mauritius',
            'Mexico',
            'Micronesia',
            'Moldova',
            'Monaco',
            'Mongolia',
            'Montenegro',
            'Morocco',
            'Mozambique',
            'Myanmar',
            'Nagorno-Karabakh',
            'Namibia',
            'Nauru',
            'Nepal',
            'Netherlands',
            'New Zealand',
            'Nicaragua',
            'Niger',
            'Nigeria',
            'Niue',
            'Northern Cyprus',
            'North Macedonia',
            'Norway',
            'Oman',
            'Pakistan',
            'Palau',
            'Palestine',
            'Panama',
            'Papua New Guinea',
            'Paraguay',
            'People\'s Republic of Korea',
            'Peru',
            'Philippines',
            'Poland',
            'Portugal',
            'Pridnestrovie',
            'Puntland',
            'Qatar',
            'Romania',
            'Russia',
            'Rwanda',
            'Saint Christopher and Nevis',
            'Saint Lucia',
            'Saint Vincent and the Grenadines',
            'Samoa',
            'San Marino',
            'São Tomé and Príncipe',
            'Saudi Arabia',
            'Senegal',
            'Serbia',
            'Seychelles',
            'Sierra Leone',
            'Singapore',
            'Slovakia',
            'Slovenia',
            'Solomon Islands',
            'Somali',
            'Somaliland',
            'South Africa',
            'South Ossetia',
            'South Sudan',
            'Spain',
            'Sri Lanka',
            'Sudan',
            'Suriname',
            'Swaziland',
            'Sweden',
            'Switzerland',
            'Syria',
            'Tajikistan',
            'Tanzania',
            'Thailand',
            'Timor-Leste',
            'Togo',
            'Tonga',
            'Trinidad and Tobago',
            'Tunisia',
            'Turkey',
            'Turkmenistan',
            'Tuvalu',
            'Uganda',
            'Ukraine',
            'United Arab Emirates',
            'United Kingdom',
            'United States',
            'Uruguay',
            'Uzbekistan',
            'Vanuatu',
            'Vatican city(the Holy see)',
            'Venezuela',
            'Vietnam',
            'Western Sahara',
            'Yemen',
            'Zambia',
            'Zimbabwe'
        )
# WOS主题词列表
all_WOS_category = (
        'Acoustics',
        'Agricultural Economics & Policy',
        'Agricultural Engineering',
        'Agriculture, Dairy & Animal Science',
        'Agriculture, Multidisciplinary',
        'Agronomy',
        'Allergy',
        'Anatomy & Morphology',
        'Andrology',
        'Anesthesiology',
        'Anthropology',
        'Archaeology',
        'Architecture',
        'Area Studies',
        'Art',
        'Asian Studies',
        'Astronomy & Astrophysics',
        'Audiology & Speech-Language Pathology',
        'Automation & Control Systems',
        'Behavioral Sciences',
        'Biochemical Research Methods',
        'Biochemistry & Molecular Biology',
        'Biodiversity Conservation',
        'Biology',
        'Biophysics',
        'Biotechnology & Applied Microbiology',
        'Business',
        'Business, Finance',
        'Cardiac & Cardiovascular Systems',
        'Cell & Tissue Engineering',
        'Cell Biology',
        'Chemistry, Analytical',
        'Chemistry, Applied',
        'Chemistry, Inorganic & Nuclear',
        'Chemistry, Medicinal',
        'Chemistry, Multidisciplinary',
        'Chemistry, Organic',
        'Chemistry, Physical',
        'Classics',
        'Clinical Neurology',
        'Communication',
        'Computer Science, Artificial Intelligence',
        'Computer Science, Cybernetics',
        'Computer Science, Hardware & Architecture',
        'Computer Science, Information Systems',
        'Computer Science, Interdisciplinary Applications',
        'Computer Science, Software Engineering',
        'Computer Science, Theory & Methods',
        'Construction & Building Technology',
        'Criminology & Penology',
        'Critical Care Medicine',
        'Crystallography',
        'Cultural Studies',
        'Dance',
        'Demography',
        'Dentistry, Oral Surgery & Medicine',
        'Dermatology',
        'Developmental Biology',
        'Ecology',
        'Economics',
        'Education & Educational Research',
        'Education, Scientific Disciplines',
        'Education, Special',
        'Electrochemistry',
        'Emergency Medicine',
        'Endocrinology & Metabolism',
        'Energy & Fuels',
        'Engineering, Aerospace',
        'Engineering, Biomedical',
        'Engineering, Chemical',
        'Engineering, Civil',
        'Engineering, Electrical & Electronic',
        'Engineering, Environmental',
        'Engineering, Geological',
        'Engineering, Industrial',
        'Engineering, Manufacturing',
        'Engineering, Marine',
        'Engineering, Mechanical',
        'Engineering, Multidisciplinary',
        'Engineering, Ocean',
        'Engineering, Petroleum',
        'Entomology',
        'Environmental Sciences',
        'Environmental Studies',
        'Ergonomics',
        'Ethics',
        'Ethnic Studies',
        'Evolutionary Biology',
        'Family Studies',
        'Film, Radio, Television',
        'Fisheries',
        'Folklore',
        'Food Science & Technology',
        'Forestry',
        'Gastroenterology & Hepatology',
        'Genetics & Heredity',
        'Geochemistry & Geophysics',
        'Geography',
        'Geography, Physical',
        'Geology',
        'Geosciences, Multidisciplinary',
        'Geriatrics & Gerontology',
        'Gerontology',
        'Health Care Sciences & Services',
        'Health Policy & Services',
        'Hematology',
        'History',
        'History & Philosophy of Science',
        'History of Social Sciences',
        'Horticulture',
        'Hospitality, Leisure, Sport & Tourism',
        'Humanities, Multidisciplinary',
        'Imaging Science & Photographic Technology',
        'Immunology',
        'Industrial Relations & Labor',
        'Infectious Diseases',
        'Information Science & Library Science',
        'Instruments & Instrumentation',
        'Integrative & Complementary Medicine',
        'International Relations',
        'Language & Linguistics',
        'Law',
        'Limnology',
        'Linguistics',
        'Literary Reviews',
        'Literary Theory & Criticism',
        'Literature',
        'Literature, African, Australian, Canadian',
        'Literature, American',
        'Literature, British Isles',
        'Literature, German, Dutch, Scandinavian',
        'Literature, Romance',
        'Literature, Slavic',
        'Logic',
        'Management',
        'Marine & Freshwater Biology',
        'Materials Science, Biomaterials',
        'Materials Science, Ceramics',
        'Materials Science, Characterization & Testing',
        'Materials Science, Coatings & Films',
        'Materials Science, Composites',
        'Materials Science, Multidisciplinary',
        'Materials Science, Paper & Wood',
        'Materials Science, Textiles',
        'Mathematical & Computational Biology',
        'Mathematics',
        'Mathematics, Applied',
        'Mathematics, Interdisciplinary Applications',
        'Mechanics',
        'Medical Ethics',
        'Medical Informatics',
        'Medical Laboratory Technology',
        'Medicine, General & Internal',
        'Medicine, Legal',
        'Medicine, Research & Experimental',
        'Medieval & Renaissance Studies',
        'Metallurgy & Metallurgical Engineering',
        'Meteorology & Atmospheric Sciences',
        'Microbiology',
        'Microscopy',
        'Mineralogy',
        'Mining & Mineral Processing',
        'Multidisciplinary Sciences',
        'Music',
        'Mycology',
        'Nanoscience & Nanotechnology',
        'Neuroimaging',
        'Neurosciences',
        'Nuclear Science & Technology',
        'Nursing',
        'Nutrition & Dietetics',
        'Obstetrics & Gynecology',
        'Oceanography',
        'Oncology',
        'Operations Research & Management Science',
        'Ophthalmology',
        'Optics',
        'Ornithology',
        'Orthopedics',
        'Otorhinolaryngology',
        'Paleontology',
        'Parasitology',
        'Pathology',
        'Pediatrics Vascular Disease',
        'Pharmacology & Pharmacy',
        'Philosophy',
        'Physics, Applied',
        'Physics, Atomic, Molecular & Chemical',
        'Physics, Condensed Matter',
        'Physics, Fluids & Plasmas',
        'Physics, Mathematical',
        'Physics, Multidisciplinary',
        'Physics, Nuclear',
        'Physics, Particles & Fields',
        'Physiology',
        'Planning & Development',
        'Plant Sciences',
        'Poetry',
        'Political Science',
        'Polymer Science',
        'Primary Health Care',
        'Psychiatry',
        'Psychology',
        'Psychology, Applied',
        'Psychology, Biological',
        'Psychology, Clinical',
        'Psychology, Developmental',
        'Psychology, Educational',
        'Psychology, Experimental',
        'Psychology, Mathematical',
        'Psychology, Multidisciplinary',
        'Psychology, Psychoanalysis',
        'Psychology, Social',
        'Public Administration',
        'Public, Environmental & Occupational Health',
        'Radiology, Nuclear Medicine & Medical Imaging',
        'Rehabilitation',
        'Religion',
        'Remote Sensing',
        'Reproductive Biology',
        'Respiratory System',
        'Rheumatology',
        'Robotics',
        'Social Issues',
        'Social Sciences, Biomedical',
        'Social Sciences, Interdisciplinary',
        'Social Sciences, Mathematical Methods',
        'Social Work',
        'Sociology',
        'Soil Science',
        'Spectroscopy',
        'Sport Sciences',
        'Statistics & Probability',
        'Substance Abuse',
        'Surgery',
        'Telecommunications',
        'Theater',
        'Thermodynamics',
        'Toxicology',
        'Transplantation',
        'Transportation',
        'Transportation Science & Technology',
        'Tropical Medicine',
        'Urban Studies',
        'Urology & Nephrology',
        'Veterinary Sciences',
        'Virology',
        'Water Resources',
        'Women\'s Studies',
        'Zoology',
        )
# Excel文件单元格边框格式
leftBorder = Border(left=Side(border_style='thin', color='000002'))
rightBorder = Border(right=Side(border_style='thin', color='000002'))
topBorder = Border(top=Side(border_style='thin', color='000002'))
bottomBorder = Border(bottom=Side(border_style='thin', color='000002'))
exceptTopBorder = Border(left=Side(border_style='thin', color='000002'), right=Side(border_style='thin', color='000002'), bottom=Side(border_style='thin',color='000002'))
exceptBottomBorder = Border(left=Side(border_style='thin', color='000002'), right=Side(border_style='thin', color='000002'), top=Side(border_style='thin',color='000002'))
allBorder = Border(left=Side(border_style='thin', color='000002'), right=Side(border_style='thin', color='000002'), top=Side(border_style='thin', color='000002'), bottom=Side(border_style='thin', color='000002'))


# 废案函数，可删除
def childWindow5(data: list):
    child = Toplevel()
    child.title('结果')
    child.geometry('400x300')
    text1 = Text(child)
    scroll1 = Scrollbar(child)
    scroll1.grid(column=1, row=1)
    text1.grid(column=0, row=1)

    scroll1.config(command=text1.yview())
    text1.config(yscrollcommand=scroll1.set)

    str1 = ''
    index = 0
    while index < len(data):
        str1 += data[index]['b.name'] + '&' + data[index]['c.name'] + '\t\t' + data[index]['d.name'] + '\t\t' +str(data[index]['count(*)']) + '\n'
        index+=1

    text1.insert(tkinter.INSERT, str1)


# 函数功能： 将数据写入Excel（通用）
# 输入：    name：文件名(不加'.xlsx')
#          sheet_name:sheet名称
#          data:格式参考：[{'名称'：'apple','数量':100},{'名称'：'banana','数量':50},{'名称'：'cow','数量':10}]
# 输出：    名称     数量
#          apple   100
#          banana  50
#          cow     10
def writeIntoExcel(name, sheet_name, data):
    ex_name = r'C:\Users\zsl\Desktop\知识图谱分析表\\' + name + '.xlsx'
    keylist = list(data[0].keys())
    coulum = len(data[0].keys())
    row = len(data)

    # 没有返回值
    if len(data) == 0:
        messagebox.showinfo(title='结果', message='未查询到此值')
        return

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print(ex_name + ' open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print(ex_name + ' open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print(sheet_name + ' sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print(sheet_name + ' sheet open success')

    # 获取行偏移量，按行读取，直到值为空
    startRow = 1
    while True:
        if ws1.cell(startRow, 1).value is None:
            break
        startRow += 1

    # 首行为数据名字
    index = 0
    while index < coulum:
        ws1.cell(startRow, index+1, keylist[index])
        index += 1

    irow = 0
    while irow < row:
        icoulum = 0
        while icoulum < coulum:
            ws1.cell(irow+startRow+1, icoulum+1, data[irow][keylist[icoulum]])
            icoulum += 1
        irow += 1

    wb.save(ex_name)


# 函数功能： 相似度变化时序生成，分为写入数据和生成图表(折线图)
# 输入：    相似度字典 key = 年份 , val = sin(key, key+1)
#
# 输出：    表格与折线图
def writeIntoExcelSimilarityDict(yearSimilarityDict):
    ex_name = '相似度时序变化图.xlsx'
    sheet_name = '数据以及图表'
    keylist = list(yearSimilarityDict.keys())

    # 新建工作簿
    wb = Workbook()
    wb.save(ex_name)

    ws1 = wb.create_sheet(sheet_name)
    wb.save(ex_name)
    sheetRow = 1
    # 写入数据
    for year in keylist:
        ws1.cell(sheetRow, 1, year)
        ws1.cell(sheetRow, 2, yearSimilarityDict[year])
        sheetRow += 1

    # 生成图表
    c1 = LineChart()
    c1.title = '相似度变化时序'
    c1.style = 12
    c1.y_axis.title = ' '
    c1.x_axis.title = '年份'
    c1.height = 20
    c1.width = 40
    # 引用范围
    data = Reference(ws1, min_col=2, min_row=1, max_row=len(keylist), max_col=2)
    c1.add_data(data)
    years = Reference(ws1, min_col=1, min_row=1, max_row=len(keylist),max_col=1)
    c1.set_categories(years)
    ws1.add_chart(c1, "F5")

    wb.save(ex_name)


# 废案函数，但包括饼图生成，暂留
def writeIntoExcelPieChartForXuekeZhicheng(name, leftyear, rightyear, data):
    ex_name = name + '.xlsx'
    sheet_name = leftyear+'-'+rightyear

    # 计算基础学科以及技术学科的数量，同时计算总数用于显示比例
    basicDic = {'category': '基础学科', 'number': 0}
    techDic = {'category': '技术学科', 'number': 0}

    for i_data in data:
        for i_category in basicAndTechTbl:
            if i_data['category'] == i_category[0]:
                if i_category[2] == 'basic':
                    basicDic['number'] += int(i_data['number'])
                if i_category[2] == 'tech':
                    techDic['number'] += int(i_data['number'])
    data.append(basicDic)
    data.append(techDic)
    allReferNum = basicDic['number'] + techDic['number']
    for i_data in data:
        i_data['percent'] = i_data['number'] / allReferNum

    keylist = list(data[0].keys())
    coulum = len(data[0].keys())
    row = len(data)

    # 没有返回值
    if len(data) == 0:
        messagebox.showinfo(title='结果', message='未查询到此值')
        return

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print('open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print('open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')



    # 获取行偏移量，按行读取，直到值为空
    startRow = 1
    while True:
        if ws1.cell(startRow, 1).value is None:
            break
        startRow += 1

    # 首行为数据名字
    index = 0
    while index < coulum:
        ws1.cell(startRow, index+1, keylist[index])
        index += 1

    irow = 0
    while irow < row:
        icoulum = 0
        while icoulum < coulum:
            ws1.cell(irow+startRow, icoulum+1, data[irow][keylist[icoulum]])
            icoulum += 1
        irow += 1

    # 重新行偏移量，按行读取，直到值为空
    startRow = 1
    while True:
        if ws1.cell(startRow, 1).value is None:
            break
        startRow += 1

    # 绘制饼图
    pie = PieChart()
    labels = Reference(ws1, min_col=1, max_col=1, min_row=1, max_row=startRow-1-2 )
    datas = Reference(ws1, min_col=2, max_col=2, min_row=1, max_row=startRow-1-2 )
    pie.add_data(datas)
    pie.set_categories(labels)
    pie.title='category'
    pie.style = 2
    pie.height = 30
    pie.width = 40
    ws1.add_chart(pie, "F5")

    pie2 = PieChart()
    labels2 = Reference(ws1, min_col=1, max_col=1, min_row=startRow - 2, max_row=startRow - 1)
    datas2 = Reference(ws1, min_col=2, max_col=2, min_row=startRow - 2, max_row=startRow - 1)
    pie2.add_data(datas2)
    pie2.set_categories(labels2)
    pie2.title = 'basic and tech'
    pie2.style = 2
    pie2.height = 30
    pie2.width = 40
    ws1.add_chart(pie2, "F66")

    wb.save(ex_name)


# 函数功能： 绘制学科支撑占比的面积堆积图
# 输入：    yearlist：年份列表
#          data：{'学科1':数量,'学科2':数量}
# 输出：    表格与面积堆积图
def writeIntoExcelXueKeZhiCheng(yearslist, data):
    ex_name = '学科支撑占比变化图.xlsx'
    sheet_name = '数据和图表'
    keyList = list(data.keys())

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print('open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print('open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    index = 2
    for year in yearslist:
        ws1.cell(column=1, row=index).value = year
        index += 1

    index = 2
    for key in keyList:
        ws1.cell(column=index, row=1).value = key
        Col = 2
        for ii in data[key]:
            ws1.cell(column=index, row=Col).value = ii
            Col += 1
        index += 1

    # 生成折线图
    if True:
        c1 = LineChart()
        c1.title = '学科支撑数量变化图'
        c1.x_axis.title = '年份'
        c1.y_axis.title = '所占百分比'
        c1.height = 20
        c1.width = 40

        datas = Reference(ws1, min_row=1, max_row=len(yearslist)+1, min_col=2, max_col=len(keyList)+1)
        c1.add_data(datas, titles_from_data=True)

        labels = Reference(ws1, min_col=1, min_row=2, max_row=len(yearslist)+1)
        c1.set_categories(labels)
        ws1.add_chart(c1, "Y4")

    # 生成面积图
    if True:
        a1 = AreaChart()
        a1.grouping = 'percentStacked'
        a1.title = '学科支撑占比变化图'
        a1.x_axis.title = '年份'
        a1.y_axis.title = '所占百分比'
        a1.height = 20
        a1.width = 40
        datas = Reference(ws1, min_row=1, max_row=len(yearslist)+1, min_col=2, max_col=len(keyList)+1)
        a1.add_data(datas, titles_from_data=True)
        labels = Reference(ws1, min_col=1, min_row=2, max_row=len(yearslist)+1)
        a1.set_categories(labels)

        ws1.add_chart(a1, "Y49")

    try:
        ws2 = wb['百分比数据']
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws2 = wb.create_sheet('百分比数据')
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    Row = 2
    while Row <= len(yearslist)+1:
        ws2.cell(row=Row, column=1).value = ws1.cell(row=Row, column=1).value
        Row += 1

    Col = 2
    diviedYearRowList = []
    while Col <= len(keyList)+1:
        ws2.cell(row=1, column=Col).value = ws1.cell(row=1, column=Col).value
        Col += 1

    # 区间分段
    Row = 2
    count = -1
    diviedYearRowList.append(Row)
    while ws1.cell(row=Row, column=1).value is not None:
        if count == 9:
            count = 0
            diviedYearRowList.append(Row)
        else:
            count += 1
        Row += 1
    diviedYearRowList.append(Row - 1)

    index = 0
    while index < len(diviedYearRowList)-1:
        startRow = diviedYearRowList[index]
        endRow = diviedYearRowList[index + 1]
        if index == len(diviedYearRowList)-2:
            endRow = diviedYearRowList[index + 1]+1
        Row = startRow
        sum = 0
        while Row < endRow:
            Col = 2
            while Col<=len(keyList)+1:
                sum += ws1.cell(row=Row, column=Col).value
                Col += 1
            Row += 1
        ws2.cell(row=endRow-1, column=len(keyList)+2).value = sum
        index += 1

    Row = len(yearslist)+1
    Col = len(keyList)+1+1
    while Row >= 2:
        if ws2.cell(row=Row, column=Col).value is None:
            ws2.cell(row=Row, column=Col).value = ws2.cell(row=Row+1, column=Col).value
        Row -= 1

    Row = 2
    while Row <= len(yearslist)+1:
        Col = 2
        while Col <= len(keyList) + 1:
            ws2.cell(row=Row, column=Col).value = ws1.cell(row=Row, column=Col).value/ws2.cell(row=Row, column=len(keyList)+1+1).value
            Col += 1
        Row += 1
    wb.save(ex_name)


# 函数功能： 基础学科与技术学科占比时序图，靠左一侧
# 输入：    leftYear:年份区间左侧
#          rightYear:年份区间右侧
#          basicdata:基础学科数据[{'category':'物理','percent':50%},{}]
#          techdata:技术学科数据[{'category':'计算机','percent':50%},{}]
# 输出：    时序图表
def drawXuekeZhichengShixuLeft(leftYear, rightYear, basicdata, techdata):
    ex_name = '学科支撑变化图.xlsx'
    sheet_name = '时序'

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print('open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print('open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    # 获取开始写的位置
    startRow = 1
    while ws1.cell(row=startRow+11, column=8).value is not None:
        startRow += 11

    index = 0
    while True:
        ws1.cell(row=startRow+index, column=8).border = leftBorder
        if index == 11:
            break
        index += 1

    # 绘制表格
    if True:
        ws1.cell(row=startRow + 2, column=2).border = allBorder
        ws1.cell(row=startRow + 2, column=3).border = allBorder
        ws1.cell(row=startRow + 3, column=2).border = allBorder
        ws1.cell(row=startRow + 3, column=3).border = allBorder
        ws1.cell(row=startRow + 4, column=2).border = allBorder
        ws1.cell(row=startRow + 4, column=3).border = allBorder
        ws1.cell(row=startRow + 7, column=2).border = allBorder
        ws1.cell(row=startRow + 7, column=3).border = allBorder
        ws1.cell(row=startRow + 8, column=2).border = allBorder
        ws1.cell(row=startRow + 8, column=3).border = allBorder
        ws1.cell(row=startRow + 9, column=2).border = allBorder
        ws1.cell(row=startRow + 9, column=3).border = allBorder
        ws1.cell(row=startRow + 3, column=5).border = allBorder
        ws1.cell(row=startRow + 4, column=5).border = allBorder
        ws1.cell(row=startRow + 7, column=5).border = allBorder
        ws1.cell(row=startRow + 8, column=5).border = allBorder

        ws1.cell(row=startRow + 3, column=4).border = exceptTopBorder
        ws1.cell(row=startRow + 7, column=4).border = exceptTopBorder
        ws1.cell(row=startRow + 7, column=6).border = exceptTopBorder
        ws1.cell(row=startRow + 5, column=7).border = exceptTopBorder

        ws1.cell(row=startRow + 4, column=6).border = exceptBottomBorder
        ws1.cell(row=startRow + 6, column=7).border = exceptBottomBorder
    # 填充数据
    if True:
        ws1.cell(row=startRow + 0, column=8).value = leftYear
        ws1.cell(row=startRow + 11, column=8).value = rightYear
        ws1.cell(row=startRow + 4, column=5).value = 'basic'
        ws1.cell(row=startRow + 3, column=5).value = basicdata[-1]['percent']
        ws1.cell(row=startRow + 8, column=5).value = 'tech'
        ws1.cell(row=startRow + 7, column=5).value = techdata[-1]['percent']

        ws1.cell(row=startRow + 2, column=3).value = basicdata[0]['category']
        ws1.cell(row=startRow + 2, column=2).value = basicdata[0]['percent']
        ws1.cell(row=startRow + 3, column=3).value = basicdata[1]['category']
        ws1.cell(row=startRow + 3, column=2).value = basicdata[1]['percent']
        ws1.cell(row=startRow + 4, column=3).value = basicdata[2]['category']
        ws1.cell(row=startRow + 4, column=2).value = basicdata[2]['percent']

        ws1.cell(row=startRow + 7, column=3).value = techdata[0]['category']
        ws1.cell(row=startRow + 7, column=2).value = techdata[0]['percent']
        ws1.cell(row=startRow + 8, column=3).value = techdata[1]['category']
        ws1.cell(row=startRow + 8, column=2).value = techdata[1]['percent']
        ws1.cell(row=startRow + 9, column=3).value = techdata[2]['category']
        ws1.cell(row=startRow + 9, column=2).value = techdata[2]['percent']

    wb.save(ex_name)


# 函数功能： 基础学科与技术学科占比时序图，靠右一侧
# 输入：    leftYear:年份区间左侧
#          rightYear:年份区间右侧
#          basicdata:基础学科数据[{'category':'物理','percent':50%},{}]
#          techdata:技术学科数据[{'category':'计算机','percent':50%},{}]
# 输出：    时序图表
def drawXuekeZhichengShixuRight(leftYear, rightYear, basicdata, techdata):
    ex_name = '学科支撑变化图.xlsx'
    sheet_name = '时序'

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print('open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print('open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    # 获取开始写的位置
    startRow = 1
    while ws1.cell(row=startRow + 11, column=8).value is not None:
        startRow += 11

    index = 0
    while True:
        ws1.cell(row=startRow + index, column=8).border = leftBorder
        if index == 11:
            break
        index += 1

    # 绘制表格
    if True:
        ws1.cell(row=startRow + 2, column=12).border = allBorder
        ws1.cell(row=startRow + 2, column=13).border = allBorder
        ws1.cell(row=startRow + 3, column=12).border = allBorder
        ws1.cell(row=startRow + 3, column=13).border = allBorder
        ws1.cell(row=startRow + 4, column=12).border = allBorder
        ws1.cell(row=startRow + 4, column=13).border = allBorder
        ws1.cell(row=startRow + 7, column=12).border = allBorder
        ws1.cell(row=startRow + 7, column=13).border = allBorder
        ws1.cell(row=startRow + 8, column=12).border = allBorder
        ws1.cell(row=startRow + 8, column=13).border = allBorder
        ws1.cell(row=startRow + 9, column=12).border = allBorder
        ws1.cell(row=startRow + 9, column=13).border = allBorder
        ws1.cell(row=startRow + 3, column=10).border = allBorder
        ws1.cell(row=startRow + 4, column=10).border = allBorder
        ws1.cell(row=startRow + 7, column=10).border = allBorder
        ws1.cell(row=startRow + 8, column=10).border = allBorder

        ws1.cell(row=startRow + 3, column=11).border = exceptTopBorder
        ws1.cell(row=startRow + 7, column=11).border = exceptTopBorder
        ws1.cell(row=startRow + 7, column=9).border = exceptTopBorder
        ws1.cell(row=startRow + 5, column=8).border = exceptTopBorder

        ws1.cell(row=startRow + 4, column=9).border = exceptBottomBorder
        ws1.cell(row=startRow + 6, column=8).border = exceptBottomBorder
    # 填充数据
    if True:
        ws1.cell(row=startRow + 0, column=8).value = leftYear
        ws1.cell(row=startRow + 11, column=8).value = rightYear
        ws1.cell(row=startRow + 4, column=10).value = 'basic'
        ws1.cell(row=startRow + 3, column=10).value = basicdata[-1]['percent']
        ws1.cell(row=startRow + 8, column=10).value = 'tech'
        ws1.cell(row=startRow + 7, column=10).value = techdata[-1]['percent']

        ws1.cell(row=startRow + 2, column=12).value = basicdata[0]['category']
        ws1.cell(row=startRow + 2, column=13).value = basicdata[0]['percent']
        ws1.cell(row=startRow + 3, column=12).value = basicdata[1]['category']
        ws1.cell(row=startRow + 3, column=13).value = basicdata[1]['percent']
        ws1.cell(row=startRow + 4, column=12).value = basicdata[2]['category']
        ws1.cell(row=startRow + 4, column=13).value = basicdata[2]['percent']

        ws1.cell(row=startRow + 7, column=12).value = techdata[0]['category']
        ws1.cell(row=startRow + 7, column=13).value = techdata[0]['percent']
        ws1.cell(row=startRow + 8, column=12).value = techdata[1]['category']
        ws1.cell(row=startRow + 8, column=13).value = techdata[1]['percent']
        ws1.cell(row=startRow + 9, column=12).value = techdata[2]['category']
        ws1.cell(row=startRow + 9, column=13).value = techdata[2]['percent']
    wb.save(ex_name)


# 函数功能： 生成中美对比总表
# 输入：    ChinaData  [{'WOS':'WOS主题词','number':100},{}]
#          USAData    [{'WOS':'WOS主题词','number':100},{}]
#          leftYear:年份区间左侧
#          rightYear:年份区间右侧
# 输出：    中美对比总表
def writeIntoExcel_allCompareButtonClicked(ChinaData, USAData, leftYear='', rightYear=''):
    ex_name = '中美对比总表.xlsx'
    sheet_name = '中美高产论文学科类别比较'+str(leftYear)+str(rightYear)

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print('open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print('open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    # 制表
    if True:
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['E'].width = 30
        ws1.cell(row=1, column=2).value = '中国'
        ws1.cell(row=1, column=5).value = '美国'
        ws1.cell(row=2, column=1).value = '排序'
        ws1.cell(row=2, column=2).value = 'Web of Science类别'
        ws1.cell(row=2, column=3).value = '数量'
        ws1.cell(row=2, column=4).value = '占比'
        ws1.cell(row=2, column=5).value = 'Web of Science类别'
        ws1.cell(row=2, column=6).value = '数量'
        ws1.cell(row=2, column=7).value = '占比'
        index = 0
        while index < 10:
            ws1.cell(row=3+index, column=1).value = index+1
            index += 1
        ws1.cell(row=13, column=1).value = '合计'

    # 填数据
    Row = 0
    ChinaSum = 0
    for i_data in ChinaData:
        ChinaSum += int(i_data['number'])
        if Row <= 9:
            ws1.cell(row=3 + Row, column=2).value = i_data['WOS']
            ws1.cell(row=3 + Row, column=3).value = i_data['number']
        Row += 1

    Row = 0
    USASum = 0
    for i_data in USAData:
        USASum += int(i_data['number'])
        if Row <= 9:
            ws1.cell(row=3 + Row, column=5).value = i_data['WOS']
            ws1.cell(row=3 + Row, column=6).value = i_data['number']
        Row += 1

    ws1.cell(row=13, column=3).value = "=SUM(C3:C12)"
    ws1.cell(row=13, column=6).value = "=SUM(F3:F12)"
    ws1.cell(row=13, column=4).value = "=SUM(D3:D12)"
    ws1.cell(row=13, column=7).value = "=SUM(G3:G12)"

    Row = 3
    while Row<=12:
        ws1.cell(row=Row, column=4).value = int(ws1.cell(row=Row, column=3).value)/ChinaSum
        ws1.cell(row=Row, column=7).value = int(ws1.cell(row=Row, column=6).value) / USASum
        Row+=1
    wb.save(ex_name)
    sheet_name = '中美高产论文共有学科类别比较'+str(leftYear)+str(rightYear)
    try:
        ws2 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws2 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    ChinaCommon = []
    USACommon = []

    index = 0
    while index < len(ChinaData) and index < len(USAData):
        for WCinUSA in USAData:
            if ChinaData[index]['WOS'] == WCinUSA['WOS']:
                if ChinaData[index] not in ChinaCommon:
                    ChinaCommon.append(ChinaData[index])
                if WCinUSA not in USACommon:
                    USACommon.append(WCinUSA)
        for WCinChina in ChinaData:
            if USAData[index]['WOS'] == WCinChina['WOS']:
                if WCinChina not in ChinaCommon:
                    ChinaCommon.append(WCinChina)
                if USAData[index] not in USACommon:
                    USACommon.append(USAData[index])
        index += 1

    ws2.cell(row=1, column=2).value = 'China'
    ws2.cell(row=1, column=3).value = 'USA'
    Row = 2
    while Row <= 11:
        ws2.cell(row=Row, column=1).value = ChinaCommon[Row - 2]['WOS']
        ws2.cell(row=Row, column=2).value = ChinaCommon[Row - 2]['number']
        ws2.cell(row=Row, column=3).value = USACommon[Row - 2]['number']
        Row += 1

    # 作图
    c1 = BarChart()
    c1.title = '中美高产论文共有学科类别比较'
    c1.x_axis.title = 'WOS类别'
    c1.y_axis.title = '数量'
    c1.width=40
    c1.height=30
    datas = Reference(ws2, min_row=1, max_row=11, min_col=2, max_col=3)
    labels = Reference(ws2, min_row=2, max_row=11, min_col=1)
    c1.add_data(datas, titles_from_data=True)
    c1.set_categories(labels)
    ws2.add_chart(c1, "E4")
    wb.save(ex_name)


# 函数功能： 学科权重雷达图
# 输入：    name:文件名(无.xlsx)
#          sheet_name:sheet名
#          data[{'category':category,'weight':50%}]
# 输出：    雷达图
def writeIntoExcel_createRadarForWeight(name, sheet_name, data):
    ex_name = name + '.xlsx'
    keylist = list(data[0].keys())
    coulum = len(data[0].keys())
    row = len(data)

    # 没有返回值
    if len(data) == 0:
        messagebox.showinfo(title='结果', message='未查询到此值')
        return

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print(ex_name + ' open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print(ex_name + ' open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print(sheet_name + ' sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print(sheet_name + ' sheet open success')

    startRow = 1

    # 首行为数据名字
    index = 0
    while index < coulum:
        ws1.cell(startRow, index+1, keylist[index])
        index += 1
    irow = 0
    while irow < row:
        icoulum = 0
        while icoulum < coulum:
            ws1.cell(irow+startRow+1, icoulum+1, data[irow][keylist[icoulum]])
            icoulum += 1
        irow += 1
    wb.save(ex_name)

    # 绘制雷达图
    c1 = RadarChart()
    data = Reference(ws1, min_col=2, min_row=2, max_row=row+1, max_col=2)
    c1.add_data(data)
    categorys = Reference(ws1, min_col=1, min_row=2, max_row=row+1, max_col=1)
    c1.set_categories(categorys)
    ws1.add_chart(c1, "F5")
    wb.save(ex_name)


# 函数功能： 生成学科占比桑基图
# 输出：    学科占比桑基图
def drawSankey():
    ex_name = '学科支撑占比变化图.xlsx'
    sheet_name = '百分比数据'

    wb = load_workbook(ex_name)
    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print('sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print('sheet open success')

    nodes = []
    links = []
    index = 2
    # 区间分段
    Row = 2
    count = -1
    diviedYearList = []
    diviedYearList.append(ws1.cell(row=Row, column=1).value)
    while ws1.cell(row=Row, column=1).value != None:
        if count == 9:
            count = 0
            diviedYearList.append(ws1.cell(row=Row, column=1).value)
        else:
            count += 1
        Row += 1
    diviedYearList.append(ws1.cell(row=Row - 1, column=1).value + 1)

    Col = 2
    while ws1.cell(row=1, column=Col).value != None:
        for year in diviedYearList:
            nodes.append(ws1.cell(row=1, column=Col).value + '_' + str(year))
        Col += 1

    def getLinkValue(str1: str, str2: str):
        ret = 0
        # 获取对应的列
        Col = 2
        while ws1.cell(row=1, column=Col).value != None:
            if ws1.cell(row=1, column=Col).value == str1[0:-5]:
                break
            Col += 1
        # 获取开始和结束的行
        Row = 2
        startR = 1
        endR = 1
        while ws1.cell(row=Row, column=1).value != None:
            if ws1.cell(row=Row, column=1).value == int(str1[-4:]):
                startR = Row
            if ws1.cell(row=Row, column=1).value == (int(str2[-4:]) - 1):
                endR = Row
                break
            Row += 1
        Row = startR
        while Row <= endR:
            ret += float(ws1.cell(row=Row, column=Col).value)
            Row += 1
        # print(str1, str2)
        return ret

    index = 1
    while index < len(nodes):
        if nodes[index - 1][0:-5] == nodes[index][0:-5]:
            links.append({'source': nodes[index - 1], 'target': nodes[index],
                          'value': getLinkValue(nodes[index - 1], nodes[index])})
        index += 1
    temp = []
    for node in nodes:
        temp.append({'name': node})
    nodes.clear()
    nodes = temp

    sankey = Sankey("桑基图示例", width=1200, height=600)
    sankey.add(
        "sankey",
        nodes,
        links,
        line_opacity=1.0,
        line_curve=0.5,
        line_color="source",
        is_label_show=False,
        label_pos="right",
    )
    sankey.render('学科支撑占比桑基图.html')


# 函数功能： 生成关键词词云图（通用）
# 输入：    file_name：文件名
#          wordsList:关键词列表
#          wordsValue：与关键词列表对应的值列表
def drawWordsCloud(file_name='', shape='', wordsList=[], wordsValue=[]):
    file_name = file_name + '.html'
    wordcloud = WordCloud(width=1300, height=620)
    wordcloud.add("", wordsList, wordsValue, word_size_range=[30, 100],
                  shape=shape)
    wordcloud.render(path=file_name)
    win32api.ShellExecute(0, 'open', file_name, '', '', 1)


# 函数功能： 生成面积堆叠图（通用）
# 输入：    file_name: 文件名
#          sheet_name: sheet名
#          x_data: 横轴列表，例如年份
#          data: 各项的数值 [{'name':[1, 2, 3]}, {'name1':[2, 3, 4]}...]
def excelPercentStacked(file_name='', sheet_name='', x_data=[], data=[]):
    ex_name = file_name + '.xlsx'

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print(ex_name + ' open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print(ex_name + ' open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print(sheet_name + ' sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print(sheet_name + ' sheet open success')

    x_row = 2
    for x in x_data:
        ws1.cell(row=x_row, column=1, value=x)
        x_row += 1

    idx_column = 2
    for i_data in data:
        ws1.cell(row=1, column=idx_column, value=list(i_data.keys())[0])
        i_row = 2
        for idx in i_data[list(i_data.keys())[0]]:
            ws1.cell(row=i_row, column=idx_column, value=idx)
            i_row += 1
        idx_column += 1

    # 生成面积堆叠图
    if True:
        a1 = AreaChart()
        a1.grouping = 'percentStacked'
        a1.title = '关键词占比变化图'
        a1.x_axis.title = '年份'
        a1.y_axis.title = '所占百分比'
        a1.height = 20
        a1.width = 40
        datas = Reference(ws1, min_row=1, max_row=len(x_data)+1, min_col=2, max_col=len(data)+1)
        a1.add_data(datas, titles_from_data=True)
        labels = Reference(ws1, min_col=1, min_row=2, max_row=len(x_data)+1)
        a1.set_categories(labels)

        ws1.add_chart(a1, "Y49")

    wb.save(ex_name)


# 函数功能： 生成折线图（通用）
# 输入：    file_name: 文件名
#          sheet_name: sheet名
#          data: [{key1:value1}, {key2:value2}, ...]
def excelLineChart(file_name='', sheet_name='', data=[{'key1':'value1'}], x_title='x', y_title='y'):
    ex_name = file_name + '.xlsx'

    try:
        wb = load_workbook(ex_name)
    except IOError:
        print(ex_name + ' open error , creating...')
        # 创建新的工作簿
        wb = Workbook()
        wb.save(ex_name)
    else:
        # 打开已有的工作簿
        print(ex_name + ' open success')

    try:
        ws1 = wb[sheet_name]
    except KeyError:
        print(sheet_name + ' sheet open error , creating...')
        # 创建新的工作表
        ws1 = wb.create_sheet(sheet_name)
        wb.save(ex_name)
    else:
        # 打开已有的工作表
        print(sheet_name + ' sheet open success')
    # 填数据
    x_row = 1
    for idx in data:
        ws1.cell(row=x_row, column=1, value=list(idx.keys())[0])
        ws1.cell(row=x_row, column=2, value=idx[list(idx.keys())[0]])
        x_row += 1
    # 生成图表
    c1 = LineChart()
    c1.title = '相似度变化时序'
    c1.style = 12
    c1.y_axis.title = y_title
    c1.x_axis.title = x_title
    c1.height = 20
    c1.width = 40
    # 引用范围
    data = Reference(ws1, min_col=2, min_row=1, max_row=len(data), max_col=2)
    c1.add_data(data)
    category = Reference(ws1, min_col=1, min_row=1, max_row=len(data), max_col=1)
    c1.set_categories(category)
    ws1.add_chart(c1, "F5")

    wb.save(ex_name)


def exportWindow(graph_entey):
    root = Toplevel()
    root.title('数据提取器-1')
    root.geometry('1400x600')

    db = NEO4J_DB()
    db.graph = graph_entey

    ######################################################################################row=0
    Label000 = Label(root,text="")
    Label000.grid(column=0, row=0)
    yearDataList = db.getYearList()
    yearDataList.pop()  # 将最新一年的数据排除，因为今年未结束。。。
    yearsList = []
    for yearData in yearDataList:
        yearsList.append(yearData['year'])

    ######################################################################################row=1
    if True:
        Label004 = Label(root,text="", width=10)
        Label004.grid(column=0, row=1)
        Label001 = Label(root,text="选择WOS关键词")
        Label001.grid(column=3, row=1)
        Combobox1 = ttk.Combobox(root, width=40)
        Combobox1.grid(column=4, row=1)
        data = db.searchAllWC()
        data_WC = []

        index = 0
        while index < len(data):
            data_WC.append(data[index]['b.name'])
            index += 1
        Combobox1['value'] = tuple(data_WC)

        def quickButtonClicked():
            cbox = Combobox1.get()

            quickButtonData = db.searchAllKeyword(cbox, str(10))
            if len(quickButtonData) != 0:
                writeIntoExcel(cbox, "主题词", quickButtonData)

            quickButtonData = db.searchAllOrgan(cbox, str(10))
            if len(quickButtonData) != 0:
                writeIntoExcel(cbox, "机构", quickButtonData)

            quickButtonData = db.searchAllCoorperateNation(cbox, str(5))
            if len(quickButtonData) != 0:
                writeIntoExcel(cbox, "合作论文国家", quickButtonData)
            CoorperateNationList = []
            #CoorperateNationList = [('China','USA'),('China','Australia'),('China','Germany'),('China','France'),('China','Japan')]
            data_Coo_Nation = db.searchCoorperatWithChina(cbox)
            for temp_data in data_Coo_Nation:
                tmpnation = []
                tmpnation.append('China')
                tmpnation.append(temp_data['c.name'])
                CoorperateNationList.append(tuple(tmpnation))

            for nation1,nation2 in CoorperateNationList:
                quickButtonData = db.searchCoorperateNationKeyword(nation1, nation2, cbox, str(10))
                if len(quickButtonData) != 0:
                    writeIntoExcel(cbox, "合作论文主题词", quickButtonData)
                quickButtonData = db.searchCoorperateNationOrgan(nation1, nation2, cbox, str(10))
                if len(quickButtonData) != 0:
                    writeIntoExcel(cbox, "合作论文机构", quickButtonData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        quickButton = Button(root, text="快速生成", command=quickButtonClicked, width=10)
        quickButton.grid(column=5, row=1)

        def allCompareButtonClicked():
            ChinaData = db.getWCDataByNation('China')
            USAData = db.getWCDataByNation('USA')
            writeIntoExcel_allCompareButtonClicked(ChinaData, USAData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', '中美对比总表.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        allCompareButton = Button(root, text="总表生成", command=allCompareButtonClicked, width=10)
        allCompareButton.grid(column=6, row=1)


        Label0701 = Label(root, text='按年份生成:', width=10)
        Label0701.grid(column=7, row=1)

        Combobox0801 = ttk.Combobox(root, width=5)
        Combobox0801.grid(column=8, row=1)
        Combobox0801['value'] = tuple(yearsList)

        Label0901 = Label(root, text='-->', width=5)
        Label0901.grid(column=9, row=1)

        Combobox1001 = ttk.Combobox(root, width=5)
        Combobox1001.grid(column=10, row=1)
        Combobox1001['value'] = tuple(yearsList)


        def clicked1101():
            leftYear = Combobox0801.get()
            rightYear = Combobox1001.get()
            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            ChinaData = db.getWCDataByNationByYear('China', leftYear, rightYear)
            USAData = db.getWCDataByNationByYear('USA', leftYear, rightYear)
            writeIntoExcel_allCompareButtonClicked(ChinaData, USAData, leftYear=leftYear, rightYear=rightYear)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', '中美对比总表.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')


        Button1101 = Button(root, text="按年份生成总表", command=clicked1101, width=14)
        Button1101.grid(column=11, row=1)


    ######################################################################################row=2
    Label005 = Label(root,text="")
    Label005.grid(column=0, row=2)

    ######################################################################################row=3
    if True:
        Label002 = Label(root,text="", width=10)
        Label002.grid(column=0, row=3)

        Label1 = Label(root, text="查询主题词",width=15)
        Label1.grid(column=3, row=3)

        spin1 = Spinbox(root, from_=1, to=10, width=6)
        spin1.grid(column=6, row=3)

        Label003 = Label(root,text="  显示数量:")
        Label003.grid(column=5, row=3)


        def clicked():
            spi = spin1.get()
            cbox = Combobox1.get()
            data = db.searchAllKeyword(cbox, spi)
            if len(data) != 0:
                writeIntoExcel(cbox, "主题词", data)

            # 自动打开Excel文件
            print(cbox)
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)

            messagebox.showinfo(title='结果', message='写入完成')


        Button1 = Button(root, text="查询", command=clicked, width=30)
        Button1.grid(column=4, row=3)

        Label0703 = Label(root, text='按年份生成:', width=10)
        Label0703.grid(column=7, row=3)

        Combobox0803 = ttk.Combobox(root, width=5)
        Combobox0803.grid(column=8, row=3)
        Combobox0803['value']=tuple(yearsList)

        Label0903 = Label(root, text='-->', width=5)
        Label0903.grid(column=9, row=3)

        Combobox1003 = ttk.Combobox(root, width=5)
        Combobox1003.grid(column=10, row=3)
        Combobox1003['value'] = tuple(yearsList)


        def clicked1103():
            leftYear = Combobox0803.get()
            rightYear = Combobox1003.get()
            num = spin1.get()
            WC = Combobox1.get()
            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            retData = db.searchAllKeywordByYear(WC, num, leftYear, rightYear)
            if len(retData) != 0:
                writeIntoExcel(WC, "主题词"+str(leftYear)+'-'+str(rightYear), retData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', WC + '.xlsx', '', '', 1)


        Button1103 = Button(root, text="按年份生成", command=clicked1103, width=10)
        Button1103.grid(column=11, row=3)

    ######################################################################################row=4
    Label006 = Label(root,text="")
    Label006.grid(column=0, row=4)

    ######################################################################################row=5
    if True:
        Label007 = Label(root,text="", width=10)
        Label007.grid(column=0, row=5)

        Label1 = Label(root, text="查询机构  ", width=15)
        Label1.grid(column=3, row=5)

        spin2 = Spinbox(root, from_=1, to=10, width=6)
        spin2.grid(column=6, row=5)

        Label008 = Label(root,text="  显示数量:")
        Label008.grid(column=5, row=5)

        def clicked2():
            spi = spin2.get()
            cbox = Combobox1.get()
            data = db.searchAllOrgan(cbox, spi)
            if len(data) != 0:
                writeIntoExcel(cbox, "机构", data)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        Button2 = Button(root, text="查询", command=clicked2, width=30)
        Button2.grid(column=4, row=5)

        Label0705 = Label(root, text='按年份生成:', width=10)
        Label0705.grid(column=7, row=5)

        Combobox0805 = ttk.Combobox(root, width=5)
        Combobox0805.grid(column=8, row=5)
        Combobox0805['value'] = tuple(yearsList)

        Label0905 = Label(root, text='-->', width=5)
        Label0905.grid(column=9, row=5)

        Combobox1005 = ttk.Combobox(root, width=5)
        Combobox1005.grid(column=10, row=5)
        Combobox1005['value'] = tuple(yearsList)


        def clicked1105():
            leftYear = Combobox0805.get()
            rightYear = Combobox1005.get()
            num = spin2.get()
            WC = Combobox1.get()
            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            retData = db.searchAllOrganByYear(WC, num, leftYear, rightYear)
            if len(retData) != 0:
                writeIntoExcel(WC, "机构" + str(leftYear) + '-' + str(rightYear), retData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', WC + '.xlsx', '', '', 1)


        Button1105 = Button(root, text="按年份生成", command=clicked1105, width=10)
        Button1105.grid(column=11, row=5)

    ######################################################################################row=6
    Label009 = Label(root,text="")
    Label009.grid(column=0, row=6)

    ######################################################################################row=7
    if True:
        Label010 = Label(root,text="", width=10)
        Label010.grid(column=0, row=7)

        Label1 = Label(root, text="合作论文国家",width=15)
        Label1.grid(column=3, row=7)

        spin3 = Spinbox(root, from_=1, to=5, width=6)
        spin3.grid(column=6, row=7)

        Label011 = Label(root,text="  显示数量:")
        Label011.grid(column=5, row=7)


        def clicked3():
            spi = spin3.get()
            cbox = Combobox1.get()
            data = db.searchAllCoorperateNation(cbox, spi)
            if len(data) != 0:
                writeIntoExcel(cbox, "合作论文国家", data)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')


        Button3 = Button(root, text="查询", command=clicked3, width=30)
        Button3.grid(column=4, row=7)

        Label0707 = Label(root, text='按年份生成:', width=10)
        Label0707.grid(column=7, row=7)

        Combobox0807 = ttk.Combobox(root, width=5)
        Combobox0807.grid(column=8, row=7)
        Combobox0807['value'] = tuple(yearsList)

        Label0907 = Label(root, text='-->', width=5)
        Label0907.grid(column=9, row=7)

        Combobox1007 = ttk.Combobox(root, width=5)
        Combobox1007.grid(column=10, row=7)
        Combobox1007['value'] = tuple(yearsList)


        def clicked1107():
            leftYear = Combobox0807.get()
            rightYear = Combobox1007.get()
            num = spin3.get()
            WC = Combobox1.get()
            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            retData = db.searchAllCoorperateNationByYear(WC, num, leftYear, rightYear)
            if len(retData) != 0:
                writeIntoExcel(WC, "国家" + str(leftYear) + '-' + str(rightYear), retData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', WC + '.xlsx', '', '', 1)

        Button1107 = Button(root, text="按年份生成", command=clicked1107, width=10)
        Button1107.grid(column=11, row=7)

    ######################################################################################row=8
    Label012 = Label(root,text="")
    Label012.grid(column=0, row=8)

    ######################################################################################row=9
    if True:
        Label013 = Label(root,text="", width=10)
        Label013.grid(column=0, row=9)

        Label1 = Label(root, text="合作论文主题词",width=15)
        Label1.grid(column=1, row=9)

        Combobox2 = ttk.Combobox(root, width=20)
        Combobox2.grid(column=2, row=9)
        Combobox2['values'] = nationList

        Combobox3 = ttk.Combobox(root, width=20)
        Combobox3.grid(column=3, row=9)
        Combobox3['values'] = nationList

        spin4 = Spinbox(root, from_=1, to=10, width=6)
        spin4.grid(column=6, row=9)

        Label014 = Label(root,text="  显示数量:")
        Label014.grid(column=5, row=9)


        def clicked4():
            cbox2 = Combobox2.get()
            cbox3 = Combobox3.get()
            spi = spin4.get()
            cbox = Combobox1.get()
            data = db.searchCoorperateNationKeyword(cbox2, cbox3, cbox, spi)
            if len(data) != 0:
                writeIntoExcel(cbox, "合作论文主题词", data)
            # print('写入完成')
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')


        Button4 = Button(root, text="查询", command=clicked4, width=30)
        Button4.grid(column=4, row=9)

        Label0709 = Label(root, text='按年份生成:', width=10)
        Label0709.grid(column=7, row=9)

        Combobox0809 = ttk.Combobox(root, width=5)
        Combobox0809.grid(column=8, row=9)
        Combobox0809['value'] = tuple(yearsList)

        Label0909 = Label(root, text='-->', width=5)
        Label0909.grid(column=9, row=9)

        Combobox1009 = ttk.Combobox(root, width=5)
        Combobox1009.grid(column=10, row=9)
        Combobox1009['value'] = tuple(yearsList)


        def clicked1109():
            cbox2 = Combobox2.get()
            cbox3 = Combobox3.get()
            leftYear = Combobox0809.get()
            rightYear = Combobox1009.get()
            num = spin4.get()
            WC = Combobox1.get()
            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            retData = db.searchCoorperateNationKeywordByYear(cbox2, cbox3, WC, num, leftYear, rightYear)
            if len(retData) != 0:
                writeIntoExcel(WC, "合作论文主题词" + str(leftYear) + '-' + str(rightYear), retData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', WC + '.xlsx', '', '', 1)


        Button1109 = Button(root, text="按年份生成", command=clicked1109, width=10)
        Button1109.grid(column=11, row=9)

    ######################################################################################row=10
    Label015 = Label(root,text="国家1", width=20)
    Label015.grid(column=2, row=10)

    Label016 = Label(root,text="国家2", width=20)
    Label016.grid(column=3, row=10)

    ######################################################################################row=11
    if True:
        Label017 = Label(root,text="", width=10)
        Label017.grid(column=0, row=11)
        Label1 = Label(root, text="合作论文机构",width=15)
        Label1.grid(column=1, row=11)

        Combobox4 = ttk.Combobox(root, width=20)
        Combobox4.grid(column=2, row=11)
        Combobox4['values'] = nationList

        Combobox5 = ttk.Combobox(root, width=20)
        Combobox5.grid(column=3, row=11)
        Combobox5['values'] = nationList

        spin5 = Spinbox(root, from_=1, to=10, width=6)
        spin5.grid(column=6, row=11)

        Label018 = Label(root,text="  显示数量:")
        Label018.grid(column=5, row=11)


        def clicked5():
            cbox4 = Combobox4.get()
            cbox5 = Combobox5.get()
            spi = spin5.get()
            cbox = Combobox1.get()
            data = db.searchCoorperateNationOrgan(cbox4, cbox5, cbox, spi)
            if len(data) != 0:
                writeIntoExcel(cbox, "合作论文机构", data)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', r'C:\Users\zsl\Desktop\知识图谱分析表\\' + cbox + '.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')


        Button5 = Button(root, text="查询", command=clicked5, width=30)
        Button5.grid(column=4, row=11)

        Label0711 = Label(root, text='按年份生成:', width=10)
        Label0711.grid(column=7, row=11)

        Combobox0811 = ttk.Combobox(root, width=5)
        Combobox0811.grid(column=8, row=11)
        Combobox0811['value'] = tuple(yearsList)

        Label0911 = Label(root, text='-->', width=5)
        Label0911.grid(column=9, row=11)

        Combobox1011 = ttk.Combobox(root, width=5)
        Combobox1011.grid(column=10, row=11)
        Combobox1011['value'] = tuple(yearsList)


        def clicked1111():
            cbox4 = Combobox4.get()
            cbox5 = Combobox5.get()
            leftYear = Combobox0811.get()
            rightYear = Combobox1011.get()
            num = spin5.get()
            WC = Combobox1.get()

            if rightYear < leftYear:
                messagebox.showinfo(title='输入异常', message='年份错误')
                return
            retData = db.searchCoorperateNationOrganByYear(cbox4, cbox5, WC, num, leftYear, rightYear)
            if len(retData) != 0:
                writeIntoExcel(WC, "合作论文机构" + str(leftYear) + '-' + str(rightYear), retData)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', WC + '.xlsx', '', '', 1)


        Button1111 = Button(root, text="按年份生成", command=clicked1111, width=10)
        Button1111.grid(column=11, row=11)

    ######################################################################################row=12
    Label019 = Label(root,text="国家1", width=20)
    Label019.grid(column=2, row=12)
    Label020 = Label(root,text="国家2", width=20)
    Label020.grid(column=3, row=12)

    ######################################################################################row=13-14
    Label021 = Label(root, text=" ", width=10)
    Label021.grid(column=3, row=13)
    Label022 = Label(root, text=" ", width=10)
    Label022.grid(column=3, row=14)

    ######################################################################################row=15
    divedYearsList = []  # 自动分割后的年份区间[[1999, 2000], [2000, 2008]...]
    if True:
        # Label023 = Label(root, text="高级图表生成:", width=10)
        # Label023.grid(column=2, row=15)
        Label024 = Label(root, text="相似度变化时序：", width=15)
        Label024.grid(column=3, row=15)

        def clicked6():
            yearDataList = db.getYearList()
            yearDataList.pop()
            yearDataList.pop()
            yearsList = []
            for yearData in yearDataList:
                yearsList.append(yearData['year'])
            # 权重向量字典
            yearSubjectWeightDict = {}
            # 获取每年的学科权重
            for yearData in yearDataList:
                # 新建字典，key = 主题 ， val = 论文数量（初值为0）
                yearDataDict = {}

                # 新建字典，key = 年份 ， val = 对应年份权重列表
                yearSubjectWeightDict[yearData['year']] = []

                maxSubjectNum = 0
                minSubjectNum = 0
                for subject in data_WC:
                    yearDataDict[subject] = 0

                # 获得对应年份的主题论文数量并升序排列，即最后一个最大
                WCData = db.getWCDataByYear(yearData['year'])
                for temp_data in WCData:
                    maxSubjectNum = (WCData[-1])['number']
                    minSubjectNum = (WCData[0])['number']
                    # 若当年主题数小于最大主题数，最小值设置为0
                    if len(WCData) < len(data_WC):
                        minSubjectNum = 0
                    # 将从数据库读到的值写入字典中
                    yearDataDict[temp_data['subject']] = temp_data['number']

                # 计算这一年每个主题的权重
                for subject in data_WC:
                    weight = (yearDataDict[subject] - minSubjectNum)/(maxSubjectNum - minSubjectNum)
                    yearSubjectWeightDict[yearData['year']].append(weight)

            # 相似度字典 key = 年份 , val = sin(key, key+1)//key = 2000, val = sin(2000, 2001)
            yearSimilarityDict = {}

            for yearData in yearDataList:
                yearSimilarityDict[yearData['year']] = 0
                # 向量空间模型
                a = numpy.array(yearSubjectWeightDict[yearData['year']])
                try :
                    b = numpy.array(yearSubjectWeightDict[(yearData['year']+1)])
                except KeyError:
                    break
                else :
                    pass
                yearSimilarityDict[yearData['year']] = (sum(a * b))/(numpy.sqrt(sum(numpy.square(a))) * numpy.sqrt(sum(numpy.square(b))))
            yearSimilarityDict.pop(yearDataList[-1]['year'])

            # 获取自动分割后的年份区间
            tempList = []
            tempList.append(yearsList[0])
            tempList.append(1988)
            divedYearsList.append(tuple(tempList))
            del tempList[0]

            for year in yearsList:
                if year <= 1988:
                    continue
                if year == yearsList[-2]:
                    break
                if yearSimilarityDict[year] < yearSimilarityDict[(int(year)+1)] \
                        and yearSimilarityDict[year] < yearSimilarityDict[(int(year) - 1)]:
                    tempList.append(year)
                    divedYearsList.append(tuple(tempList))
                    del tempList[0]
            tempList.append(yearsList[-1])
            divedYearsList.append(tuple(tempList))

            writeIntoExcelSimilarityDict(yearSimilarityDict)
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', '相似度时序变化图.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        Button6 = Button(root, text="生成", command=clicked6, width=30)
        Button6.grid(column=4, row=15)
        Label1506 = Label(root, text="导出知识图谱:", width=10)
        Label1506.grid(column=6, row=15)

        def clicked1508():
            db.export_lunwen_node()
            db.export_zuozhe_node()
            db.export_guojia_node()
            db.export_jigou_node()
            db.export_xueke_node()
            db.export_zuozhe_guojia_rela()
            db.export_zuozhe_jigou_rela()
            db.export_lunwen_zuozhe_rela()

        Button1508 = Button(root, text="导出", command=clicked1508, width=10)
        Button1508.grid(column=8, row=15)

    ######################################################################################row=16
    Label025 = Label(root, text=" ", width=10)
    Label025.grid(column=3, row=16)

    ######################################################################################row=17
    if True:
        Label026 = Label(root, text="学科支撑变化时序：", width=17)
        Label026.grid(column=3, row=17)

        def clicked7():
            leftflag = True
            if len(divedYearsList) == 0:
                clicked6()
            for leftYear, rightYear in divedYearsList:
                LeibieData = db.getLeibieByYear(leftYear, rightYear)
                # 计算基础学科以及技术学科的数量，同时计算总数用于显示比例
                basicDic = {'category': '基础学科', 'number': 0}
                techDic = {'category': '技术学科', 'number': 0}
                for i_data in LeibieData:
                    for i_category in basicAndTechTbl:
                        if i_data['category'] == i_category[0]:
                            if i_category[2] == 'basic':
                                basicDic['number'] += int(i_data['number'])
                            if i_category[2] == 'tech':
                                techDic['number'] += int(i_data['number'])
                LeibieData.append(basicDic)
                LeibieData.append(techDic)
                allReferNum = basicDic['number'] + techDic['number']

                for i_data in LeibieData:
                    i_data['percent'] = i_data['number'] / allReferNum

                basicDicList=[]
                techDicList=[]

                for i_data in LeibieData:
                    for i_category in basicAndTechTbl:
                        if i_data['category'] == i_category[0]:
                            if i_category[2] == 'basic':
                                basicDicList.append(i_data)
                            if i_category[2] == 'tech':
                                techDicList.append(i_data)
                basicDicList.append(basicDic)
                techDicList.append(techDic)

                if leftflag:
                    drawXuekeZhichengShixuLeft(leftYear,rightYear,basicDicList,techDicList)
                    leftflag=False

                else:
                    drawXuekeZhichengShixuRight(leftYear,rightYear,basicDicList,techDicList)
                    leftflag = True
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', '学科支撑变化图.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        Button7 = Button(root, text="生成", command=clicked7, width=30)
        Button7.grid(column=4, row=17)

        Label1706 = Label(root, text="标签与摘要", width=10)
        Label1706.grid(column=6, row=17)

        def click1708():
            tag_window = Toplevel()
            tag_window.title('标签与摘要')
            tag_window.geometry('450x200')

            Label_tag_0000 = Label(tag_window, text='', width=10)
            Label_tag_0000.grid(column=0, row=0)
            Label_tag_0101 = Label(tag_window, text="导出标签:", width=10)
            Label_tag_0101.grid(column=1, row=1)
            Label_tag_0102 = Label(tag_window, text="数量", width=10)
            Label_tag_0102.grid(column=2, row=1)
            Combobox_tag_0103 = ttk.Combobox(tag_window, width=10)
            Combobox_tag_0103.grid(column=3, row=1)
            Combobox_tag_0103['value'] = (10, 20, 50, 100, 1000, 10000)

            def click_Button_tag_0104():
                num = Combobox_tag_0103.get()
                data = db.getAllTag(num)
                writeIntoExcel('全标签', 'TOP'+str(num), data)
                # 自动打开Excel文件
                win32api.ShellExecute(0, 'open', '全标签.xlsx', '', '', 1)

            Button_tag_0104 = Button(tag_window, text='导出', command=click_Button_tag_0104, width=10)
            Button_tag_0104.grid(row=1, column=4)

            Label_tag_0201 = Label(tag_window, text="作者标签及摘要:", width=15)
            Label_tag_0201.grid(row=2, column=1)
            Entry_tag_0202 = Entry(tag_window, width=10)
            Entry_tag_0202.grid(row=2, column=2)

            def click_Button_tag_0203():
                author = Entry_tag_0202.get()
                data1 = db.getTagByAuthor(author)
                writeIntoExcel('作者标签及摘要', author, data1)
                data2 = db.getAbstractByAuthor(author)
                writeIntoExcel('作者标签及摘要', author, data2)
                # 自动打开Excel文件
                win32api.ShellExecute(0, 'open', '作者标签及摘要.xlsx', '', '', 1)

            Button_tag_0203 = Button(tag_window, text='导出', command=click_Button_tag_0203, width=10)
            Button_tag_0203.grid(row=2, column=3)

        Button1708 = Button(root, text="打开",command=click1708 , width=10)
        Button1708.grid(column=8, row=17)

    ######################################################################################row=18
    Label027 = Label(root, text=" ", width=10)
    Label027.grid(column=3, row=18)

    ######################################################################################row=19
    if True:
        Label028 = Label(root, text="学科支撑比例时序：", width=19)
        Label028.grid(column=3, row=19)

        def clicked9():
            # 获取年份数值列表
            yearDataList = db.getYearList()
            yearDataList.pop()
            yearDataList.pop()
            yearsList = []

            for yearData in yearDataList:
                yearsList.append(yearData['year'])
            ReferSOTypeData = db.getReferSOType()
            ReferSOTypeList =[]

            for i_data in ReferSOTypeData:
                ReferSOTypeList.append(i_data['category'])
            # 定义储存数据的字典
            XueKeZhiChengDict = {}

            for category in ReferSOTypeList:
                XueKeZhiChengDict[category] = []
            maxLen = 0

            for year in yearsList:
                # 获取某年的引用期刊类别数据
                yearData = db.getLeibieByYear(year, year)
                for i_data in yearData:
                    XueKeZhiChengDict[i_data['category']].append(i_data['number'])
                    if len(XueKeZhiChengDict[i_data['category']]) >= maxLen:
                        maxLen = len(XueKeZhiChengDict[i_data['category']])
                for key in list(XueKeZhiChengDict.keys()):
                    if len(XueKeZhiChengDict[key]) < maxLen:
                        XueKeZhiChengDict[key].append(0)

            writeIntoExcelXueKeZhiCheng(yearsList, XueKeZhiChengDict)
            drawSankey()
            # 自动打开Excel文件
            win32api.ShellExecute(0, 'open', '学科支撑占比变化图.xlsx', '', '', 1)
            messagebox.showinfo(title='结果', message='写入完成')

        Button9 = Button(root, text="生成", command=clicked9, width=30)
        Button9.grid(column=4, row=19)

        Label1906 = Label(root, text="关键词突变", width=10)
        Label1906.grid(column=6, row=19)

        def click1908():
            keyChange_window = Toplevel()
            keyChange_window.title('关键词突变情况')
            keyChange_window.geometry('600x200')

            Label_tag_0000 = Label(keyChange_window, text='', width=10)
            Label_tag_0000.grid(column=0, row=0)
            Label_tag_0101 = Label(keyChange_window, text="选择年份", width=10)
            Label_tag_0101.grid(column=1, row=1)
            Combobox_tag_0102 = ttk.Combobox(keyChange_window, width=10)
            Combobox_tag_0102.grid(column=2, row=1)
            Combobox_tag_0102['value'] = tuple(yearsList)
            Label_tag_0103 = Label(keyChange_window, text="-->", width=5)
            Label_tag_0103.grid(column=3, row=1)
            Combobox_tag_0104 = ttk.Combobox(keyChange_window, width=10)
            Combobox_tag_0104.grid(column=4, row=1)
            Combobox_tag_0104['value'] = tuple(yearsList)

            def click_Button_tag_0104():
                left_year = Combobox_tag_0102.get()
                right_year = Combobox_tag_0104.get()

                # 获取两个年份的关键词及数量[{'关键词':'xxx','数量':xxx},{}]
                left_year_allKeyword = db.searchAllKeywordOnlyByYear(num='1000', leftY=left_year, rightY=left_year)
                right_year_allKeyword = db.searchAllKeywordOnlyByYear(num='1000', leftY=right_year, rightY=right_year)

                # 将两个年份的关键词列表“对齐”
                leftList = left_year_allKeyword
                rightList = []
                for leftData in left_year_allKeyword:
                    rightList.append({'关键词':leftData['关键词'],'数量':0})
                    for rightData in right_year_allKeyword:
                        if leftData['关键词'] == rightData['关键词']:
                            rightList[-1]['数量'] = rightData['数量']

                left_year_paper_num = db.countPaperByYear(leftY=left_year, rightY=left_year)
                right_year_paper_num = db.countPaperByYear(leftY=right_year, rightY=right_year)

                # 根据a,b,c,d的值计算LL值并返回
                def keywordsFrequency(a, b, c, d):
                    if a == 0 and b == 0:
                        return 0
                    if a != 0 and b != 0:
                        e1 = a * math.log(a / (c * (a + b) / (c + d)))
                        e2 = b * math.log(b / (d * (a + b) / (c + d)))
                        LL = 2 * (e1 + e2)
                        if e2 < 0:
                            LL = -LL
                        LL = round(LL, 3)
                        return LL
                    elif a == 0 and b != 0:
                        e2 = b * math.log(b / (d * (a + b) / (c + d)))
                        LL = 2 * e2
                        LL = round(LL, 3)
                        return LL
                    elif a != 0 and b == 0:
                        e1 = a * math.log(a / (c * (a + b) / (c + d)))
                        LL = 2 * e1
                        LL = round(LL, 3)
                        return LL

                # [{'关键词':'xxx','LL值': xxx},{},{}]
                LL_List = []
                index = 0
                while index < len(leftList):
                    LL = keywordsFrequency(leftList[index]['数量'], rightList[index]['数量'], left_year_paper_num[0]['数量'], right_year_paper_num[0]['数量'])
                    LL_List.append({'关键词':leftList[index]['关键词'], '前一年':leftList[index]['数量'],'后一年':rightList[index]['数量'], 'LL值': LL})
                    index += 1
                # 写入Excel并自动打开
                writeIntoExcel('关键词变化', str(left_year)+'-'+str(right_year), LL_List)
                win32api.ShellExecute(0, 'open', '关键词变化.xlsx', '', '', 1)

            Button_tag_0104 = Button(keyChange_window, text='导出', command=click_Button_tag_0104, width=10)
            Button_tag_0104.grid(row=1, column=5)


        Button1908 = Button(root, text="打开", command=click1908, width=10)
        Button1908.grid(column=8, row=19)

    ######################################################################################row=20
    Label2003 = Label(root, text=" ", width=10)
    Label2003.grid(column=3, row=20)

    ######################################################################################row=21
    if True:
        Label2103 = Label(root, text="引文学科权重图：", width=21)
        Label2103.grid(column=3, row=21)

        def clicked2104():
            top1 = Toplevel()
            top1.title('学科权重图')
            top1.geometry('450x200')

            Label_top1_0000 = Label(top1, text=" ", width=10)
            Label_top1_0000.grid(column=0, row=0)
            Label_top1_0101 = Label(top1, text="年份:", width=10)
            Label_top1_0101.grid(column=1, row=1)

            Combobox_top1_0102 = ttk.Combobox(top1, width=10)
            Combobox_top1_0102.grid(column=2, row=1)
            Combobox_top1_0102['value'] = tuple(yearsList)

            Label_top1_0103 = Label(top1, text='-->', width=5)
            Label_top1_0103.grid(column=3, row=1)

            Combobox_top1_0104 = ttk.Combobox(top1, width=10)
            Combobox_top1_0104.grid(column=4, row=1)
            Combobox_top1_0104['value'] = tuple(yearsList)

            Label_top1_0201 = Label(top1, text="国家:", width=10)
            Label_top1_0201.grid(column=1, row=2)

            Combobox_top1_0202 = ttk.Combobox(top1, width=10)
            Combobox_top1_0202.grid(column=2, row=2)
            Combobox_top1_0202['values'] = nationList

            Label_top1_0301 = Label(top1, text="生成:", width=10)
            Label_top1_0301.grid(column=1, row=3)

            def clicked_top1_0302():
                leftYear = Combobox_top1_0102.get()
                rightYear = Combobox_top1_0104.get()
                nation = Combobox_top1_0202.get()

                data1 = db.getLeibieByYearOrderByName(leftYear, rightYear)
                data2 = db.getLeibieByNationByYearOrderByName(nation, leftYear, rightYear)
                Wb = 0
                for data in data2:
                    Wb += int(data['number'])

                categoryList = []
                for i in data1:
                    categoryList.append(i['category'])
                Wa = 0
                p = []
                for category in categoryList:
                    for data_1 in data1:
                        if category == data_1['category']:
                            Wa = int(data_1['number'])
                    for data_2 in data2:
                        if category == data_2['category']:
                            Wi = int(data_2['number'])
                            p.append({'category':category,'weight':sqrt(float((Wi/Wa)*(Wi/Wa)) + float((Wi/Wb)*(Wi/Wb)))})

                writeIntoExcel_createRadarForWeight('引文学科权重图', nation+str(leftYear)+'-'+str(rightYear), p)
                # 自动打开Excel文件
                win32api.ShellExecute(0, 'open', '引文学科权重图.xlsx', '', '', 1)

            Button_top1_0302 = Button(top1, text="生成", command=clicked_top1_0302, width=10)
            Button_top1_0302.grid(column=2, row=3)

        Button2104 = Button(root, text="选择年份和国家", command=clicked2104, width=30)
        Button2104.grid(column=4, row=21)

        Label2106 = Label(root, text="jaro关键词合并", width=12)
        Label2106.grid(column=6, row=21)

        def click2108():
            data = db.getKeywodsName()
            KeywordList = []
            # [变化前， 变化后]
            ChangeList = []

            # 将数据库的关键词存入列表
            for idx in data:
                KeywordList.append(idx['keyword'])
            # print(KeywordList)

            def jaroSim(word1, word2):
                """
                计算jaro距离。
                :param word1: 词
                :param word2: 词
                :return: jaro距离
                """
                return Levenshtein.jaro(word1, word2)

            i = 0
            while i < len(KeywordList):
                j = i + 1
                # print('i : ', i, ',', 'j : ', j)
                while j < len(KeywordList):
                    if jaroSim(KeywordList[i], KeywordList[j]) >= 0.9:
                        ChangeList.append((KeywordList[j], KeywordList[i]))
                        del KeywordList[j]
                        j -= 1
                    j += 1
                i += 1
            # print(ChangeList)

            for before, after in ChangeList:
                db.changeKeyword(before=before, after=after)

        Button2108 = Button(root, text="运行", command=click2108, width=10)
        Button2108.grid(column=8, row=21)

        def click2110():
            Page_2_Window = Toplevel()
            Page_2_Window.title('数据提取器-2')
            Page_2_Window.geometry('1400x600')

            # page 2 : row = 0
            Label_page2_0000 = Label(Page_2_Window, text='', width=10)
            Label_page2_0000.grid(row=0, column=0)

            # page 2 : row = 1
            Label_page2_0100 = Label(Page_2_Window, text='', width=10)
            Label_page2_0100.grid(row=1, column=0)

            Label_page2_0101 = Label(Page_2_Window, text='关键词词云', width=10)
            Label_page2_0101.grid(row=1, column=1)

            def click_Button_page2_0102():
                keywordCloudWindow = Toplevel()
                keywordCloudWindow.title('关键词云')
                keywordCloudWindow.geometry('450x200')

                Label_keycloud_0000 = Label(keywordCloudWindow, text='', width=10)
                Label_keycloud_0000.grid(row=0, column=0)
                Label_keycloud_0100 = Label(keywordCloudWindow, text='词云类型:', width=10)
                Label_keycloud_0100.grid(row=1, column=0)
                Combobox_keycloud_0101 = ttk.Combobox(keywordCloudWindow, width=10)
                Combobox_keycloud_0101.grid(row=1, column=1)
                Combobox_keycloud_0101['value']=("cardioid", "diamond", "triangle-forward", "triangle", "pentagon", "star")
                Label_keycloud_0102 = Label(keywordCloudWindow, text='TOP', width=10)
                Label_keycloud_0102.grid(row=1, column=2)
                Combobox_keycloud_0103 = ttk.Combobox(keywordCloudWindow, width=10)
                Combobox_keycloud_0103.grid(row=1, column=3)
                Combobox_keycloud_0103['value'] = (10, 20, 30, 50, 100)

                def click_Button_keyCloud_0104():
                    type = Combobox_keycloud_0101.get()
                    number = Combobox_keycloud_0103.get()
                    data = db.getKeywodsNameAndCount(limit_num=number)
                    keyword = []
                    valuelist = []
                    for idx in data:
                        keyword.append(idx['关键词'])
                        valuelist.append((idx['数量']))
                    drawWordsCloud(file_name='关键词词云', shape=type, wordsList=keyword, wordsValue=valuelist)

                Button_keyCloud_0104 = Button(keywordCloudWindow, text="生成", command=click_Button_keyCloud_0104, width=10)
                Button_keyCloud_0104.grid(row=1, column=4)

            Button_page2_0102 = Button(Page_2_Window, text="选择参数", command=click_Button_page2_0102, width=10)
            Button_page2_0102.grid(row=1, column=2)

            Label_page2_0103 = Label(Page_2_Window, text='', width=5)
            Label_page2_0103.grid(row=1, column=3)

            Label_page2_0104 = Label(Page_2_Window, text='关键词占比', width=10)
            Label_page2_0104.grid(row=1, column=4)

            def click_Button_page2_0105():
                keywordPercentWindow = Toplevel()
                keywordPercentWindow.title('关键词占比')
                keywordPercentWindow.geometry('500x200')

                Label_keyPercent_0000 = Label(keywordPercentWindow, text='', width=10)
                Label_keyPercent_0000.grid(row=0, column=0)
                Label_keyPercent_0100 = Label(keywordPercentWindow, text='选择年份', width=10)
                Label_keyPercent_0100.grid(row=1, column=0)
                Combobox_keyPercent_0101 = ttk.Combobox(keywordPercentWindow, width=10)
                Combobox_keyPercent_0101.grid(row=1, column=1)
                Combobox_keyPercent_0101['value'] = tuple(yearsList)
                Label_keyPercent_0102 = Label(keywordPercentWindow, text='--', width=5)
                Label_keyPercent_0102.grid(row=1, column=2)
                Combobox_keyPercent_0103 = ttk.Combobox(keywordPercentWindow, width=10)
                Combobox_keyPercent_0103.grid(row=1, column=3)
                Combobox_keyPercent_0103['value'] = tuple(yearsList)
                Spinbox_keyPercent_0105 = Spinbox(keywordPercentWindow, from_=1, to=100, width=10)
                Spinbox_keyPercent_0105.grid(row=1, column=5)

                def click_Button_keyPercent_0104():
                    leftYear = Combobox_keyPercent_0101.get()
                    rightYear = Combobox_keyPercent_0103.get()
                    limit_num = int(Spinbox_keyPercent_0105.get())
                    excelData = []
                    # 先获取年份区间内的关键词TOP100列表
                    data = db.searchAllKeywordOnlyByYear(leftY=leftYear, rightY=rightYear, num=limit_num)
                    keywordsList=[]
                    for idx in data:
                        keywordsList.append(idx['关键词'])
                    # 获取这前100关键词每年的数量
                    for word in keywordsList:
                        dic = db.getNumberByKeywordEveryYear(leftY=leftYear, rightY=rightYear, keyword=word)
                        excelData.append(dic)
                    # 获取年份列表
                    year_list = []
                    year = int(leftYear)
                    while year <= int(rightYear):
                        year_list.append(year)
                        year += 1
                    # 写入excel
                    excelPercentStacked(file_name='关键词占比面积图', sheet_name=str(leftYear)+'-'+str(rightYear), x_data=year_list, data=excelData)
                    win32api.ShellExecute(0, 'open', '关键词占比面积图.xlsx', '', '', 1)

                Button_keyPercent_0104 = Button(keywordPercentWindow, text='生成', command=click_Button_keyPercent_0104, width=10)
                Button_keyPercent_0104.grid(row=1, column=4)

            Button_page2_0105 = Button(Page_2_Window, text="选择年份", command=click_Button_page2_0105, width=10)
            Button_page2_0105.grid(row=1, column=5)

        Button2108 = Button(root, text="下一页", command=click2110, width=10)
        Button2108.grid(column=10, row=21)

    root.mainloop()


if __name__ == '__main__':
    baseWindow = Tk()
    baseWindow.title('工程科学知识图谱构建与分析系统')
    baseWindow.geometry('600x300')

    # baseWindow row = 0
    Label_base_0000 = Label(baseWindow, text="  ", width=10)
    Label_base_0000.grid(row=0, column=0)

    # baseWindow row = 1
    Label_base_0100 = Label(baseWindow, text="  ", width=10)
    Label_base_0100.grid(row=1, column=0)

    # baseWindow row = 2
    Label_base_0200 = Label(baseWindow, text="  ", width=30)
    Label_base_0200.grid(row=2, column=0)

    def click_Button_base_0201():
        down_infor_window = Toplevel()
        down_infor_window.title('获取数据库信息')
        down_infor_window.geometry('450x200')

        # DBinformation row = 0
        Label_download_0000 = Label(down_infor_window, text="  ", width=10)
        Label_download_0000.grid(row=0, column=0)

        # DBinformation row = 1
        Label_download_0100 = Label(down_infor_window, text="  ", width=15)
        Label_download_0100.grid(row=1, column=0)

        Label_download_0101 = Label(down_infor_window, text="期刊列表路径: ", width=15)
        Label_download_0101.grid(row=1, column=1)

        Entry_download_0102 = Entry(down_infor_window, width=20)
        Entry_download_0102.grid(row=1, column=2)

        # DBinformation row = 2
        Label_download_0200 = Label(down_infor_window, text="  ", width=15)
        Label_download_0200.grid(row=2, column=0)

        Label_download_0201 = Label(down_infor_window, text="下载保存路径: ", width=15)
        Label_download_0201.grid(row=2, column=1)

        Entry_download_0202 = Entry(down_infor_window, width=20)
        Entry_download_0202.grid(row=2, column=2)

        # DBinformation row = 3
        Label_download_0300 = Label(down_infor_window, text="  ", width=15)
        Label_download_0300.grid(row=3, column=0)

        Label_download_0301 = Label(down_infor_window, text="SID: ", width=15)
        Label_download_0301.grid(row=3, column=1)

        Entry_download_0302 = Entry(down_infor_window, width=20)
        Entry_download_0302.grid(row=3, column=2)

        def click_Button_download_0402():
            journalListPath = Entry_download_0102.get()
            outputPath = Entry_download_0202.get()
            SID = Entry_download_0302.get()
            run_crawl_process(journalListPath, outputPath, SID)

        Button_download_0402 = Button(down_infor_window, text="开始下载", command=click_Button_download_0402, width=20)
        Button_download_0402.grid(row=4, column=2)

    Button_base_0201 = Button(baseWindow, text="下载论文", command=click_Button_base_0201, width=20)
    Button_base_0201.grid(row=2, column=1)

    # baseWindow row = 3
    Label_base_0300 = Label(baseWindow, text="  ", width=10)
    Label_base_0300.grid(row=3, column=0)

    # baseWindow row = 4
    Label_base_0400 = Label(baseWindow, text="  ", width=30)
    Label_base_0400.grid(row=4, column=0)

    def click_Button_base_0401():
        get_DBinformation_window = Toplevel()
        get_DBinformation_window.title('获取数据库信息')
        get_DBinformation_window.geometry('450x200')

        # DBinformation row = 0
        Label_dbinfor_0000 = Label(get_DBinformation_window, text="  ", width=10)
        Label_dbinfor_0000.grid(row=0, column=0)

        # DBinformation row = 1
        Label_dbinfor_0100 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0100.grid(row=1, column=0)

        Label_dbinfor_0101 = Label(get_DBinformation_window, text="http port: ", width=10)
        Label_dbinfor_0101.grid(row=1, column=1)

        Entry_dbinfor_0102 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0102.grid(row=1, column=2)

        # DBinformation row = 2
        Label_dbinfor_0200 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0200.grid(row=2, column=0)

        Label_dbinfor_0201 = Label(get_DBinformation_window, text="username: ", width=10)
        Label_dbinfor_0201.grid(row=2, column=1)

        Entry_dbinfor_0202 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0202.grid(row=2, column=2)

        # DBinformation row = 3
        Label_dbinfor_0300 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0300.grid(row=3, column=0)

        Label_dbinfor_0301 = Label(get_DBinformation_window, text="password: ", width=10)
        Label_dbinfor_0301.grid(row=3, column=1)

        Entry_dbinfor_0302 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0302.grid(row=3, column=2)

        # DBinformation row = 4

        def click_Button_dbinfor_0402():
            http_infor = Entry_dbinfor_0102.get()
            usrname_infor = Entry_dbinfor_0202.get()
            password_infor = Entry_dbinfor_0302.get()
            graph_entey = Graph('http://localhost:'+http_infor, username=usrname_infor, password=password_infor)

            get_DBinformation_window.destroy()

            import_To_database = Toplevel()
            import_To_database.title('导入数据库')
            import_To_database.geometry('450x200')

            # import row = 0
            Label_import_0000 = Label(import_To_database, text="  ", width=10)
            Label_import_0000.grid(row=0, column=0)

            # import row = 1
            Label_import_0100 = Label(import_To_database, text="  ", width=15)
            Label_import_0100.grid(row=1, column=0)

            Label_import_0101 = Label(import_To_database, text="论文存放路径: ", width=10)
            Label_import_0101.grid(row=1, column=1)

            Entry_import_0102 = Entry(import_To_database, width=20)
            Entry_import_0102.grid(row=1, column=2)

            # import row = 3
            Label_import_0300 = Label(import_To_database, text="  ", width=15)
            Label_import_0300.grid(row=3, column=0)

            # import row = 4
            Label_import_0401 = Label(import_To_database, text="停用词表路径: ", width=10)
            Label_import_0401.grid(row=4, column=1)

            Entry_import_0402 = Entry(import_To_database, width=20)
            Entry_import_0402.grid(row=4, column=2)

            def click_Button_import_0202():
                getAllTxtData(Entry_import_0102.get(), graph_entey, Entry_import_0402.get())

            Button_import_0202 = Button(import_To_database, text="导入", command=click_Button_import_0202, width=20)
            Button_import_0202.grid(row=2, column=2)

        Button_dbinfor_0402 = Button(get_DBinformation_window, text="确认", command=click_Button_dbinfor_0402, width=10)
        Button_dbinfor_0402.grid(row=4, column=2)

        # end : click_Button_base_0401

    Button_base_0401 = Button(baseWindow, text="导入数据", command=click_Button_base_0401, width=20)
    Button_base_0401.grid(row=4, column=1)

    # baseWindow row = 5
    Label_base_0500 = Label(baseWindow, text="  ", width=10)
    Label_base_0500.grid(row=5, column=0)

    # baseWindow row = 6
    Label_base_0600 = Label(baseWindow, text="  ", width=30)
    Label_base_0600.grid(row=6, column=0)

    def click_Button_base_0601():
        get_DBinformation_window = Toplevel()
        get_DBinformation_window.title('获取数据库信息')
        get_DBinformation_window.geometry('450x200')

        # DBinformation row = 0
        Label_dbinfor_0000 = Label(get_DBinformation_window, text="  ", width=10)
        Label_dbinfor_0000.grid(row=0, column=0)

        # DBinformation row = 1
        Label_dbinfor_0100 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0100.grid(row=1, column=0)

        Label_dbinfor_0101 = Label(get_DBinformation_window, text="http port: ", width=10)
        Label_dbinfor_0101.grid(row=1, column=1)

        Entry_dbinfor_0102 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0102.grid(row=1, column=2)

        # DBinformation row = 2
        Label_dbinfor_0200 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0200.grid(row=2, column=0)

        Label_dbinfor_0201 = Label(get_DBinformation_window, text="username: ", width=10)
        Label_dbinfor_0201.grid(row=2, column=1)

        Entry_dbinfor_0202 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0202.grid(row=2, column=2)

        # DBinformation row = 3
        Label_dbinfor_0300 = Label(get_DBinformation_window, text="  ", width=15)
        Label_dbinfor_0300.grid(row=3, column=0)

        Label_dbinfor_0301 = Label(get_DBinformation_window, text="password: ", width=10)
        Label_dbinfor_0301.grid(row=3, column=1)

        Entry_dbinfor_0302 = Entry(get_DBinformation_window, width=10)
        Entry_dbinfor_0302.grid(row=3, column=2)

        # DBinformation row = 4

        def click_Button_dbinfor_0402():
            http_infor = Entry_dbinfor_0102.get()
            usrname_infor = Entry_dbinfor_0202.get()
            password_infor = Entry_dbinfor_0302.get()
            graph_entey = Graph('http://localhost:' + http_infor, username=usrname_infor, password=password_infor)

            get_DBinformation_window.destroy()
            exportWindow(graph_entey)

        Button_dbinfor_0402 = Button(get_DBinformation_window, text="确认", command=click_Button_dbinfor_0402, width=10)
        Button_dbinfor_0402.grid(row=4, column=2)

    Button_base_0601 = Button(baseWindow, text="图表生成", command=click_Button_base_0601, width=20)
    Button_base_0601.grid(row=6, column=1)


    baseWindow.mainloop()